import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = '/Users/brandonroy/Desktop/Informatique/Projects/VIVC-Python'
DATA = f'{BASE}/data'

# ─── Palette ─────────────────────────────────────────────────────────────────
C = dict(
    hdr_bg    = "4A235A",
    hdr_fg    = "FFFFFF",
    match     = "C6EFCE",   # green
    mismatch  = "FFC7CE",   # red
    single    = "FFEB9C",   # amber – only one source
    absent    = "F2F2F2",   # grey
    anchor    = "EAE4F2",   # lavender
    genre_hdr = "1A237E",   # dark blue – GENRES081 loci
    extra_hdr = "4E342E",   # dark brown – extra loci
    src_vivc  = "6A1B9A",
    src_lac   = "1B5E20",
    src_mig   = "0D47A1",
    src_oli   = "B71C1C",
    src_vas   = "880E4F",
    src_ran   = "E65100",
    src_tel   = "1565A0",   # steel blue – Tello 2024
    row_alt   = "F8F6FB",
)

THIN  = Side(style="thin",   color="C0C0C0")
MED   = Side(style="medium", color="888888")
bdr   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
bdr_m = Border(left=MED,  right=THIN, top=THIN, bottom=THIN)

def hfill(hex_): return PatternFill("solid", fgColor=hex_)
def hfont(sz=10, bold=True, color="FFFFFF"): return Font(name="Arial", bold=bold, size=sz, color=color)
def bfont(sz=9, bold=False, color="000000"): return Font(name="Arial", size=sz, bold=bold, color=color)
def cal(h="center", v="center", w=False): return Alignment(horizontal=h, vertical=v, wrap_text=w)

# ─── Load passport (has scraping artifact header row) ────────────────────────
passport_raw = pd.read_csv(f'{BASE}/vivc_passport_table.csv', low_memory=False)
# Row 0 is the real header
passport_raw.columns = passport_raw.iloc[0].tolist()
passport_raw = passport_raw.iloc[1:].reset_index(drop=True)
passport = passport_raw[['Prime name','Variety number VIVC']].copy()
passport.columns = ['vivc_prime_name','vivc_no']
passport['vivc_no'] = passport['vivc_no'].astype(str).str.strip().str.split('.').str[0]
no2name = dict(zip(passport['vivc_no'], passport['vivc_prime_name'].str.upper()))

datasrc = pd.read_csv(f'{DATA}/datasource_inventory.csv', low_memory=False)

# ─── Locus sets ──────────────────────────────────────────────────────────────
GENRES = ['VVS2','VVMD5','VVMD7','VVMD25','VVMD27','VVMD28','VVMD32','VrZAG62','VrZAG79']
LAC_EXTRA = ['VVMD21','VVMD24','VMC1b11','VMC4f3','VVIb01','VVIh54',
             'VVIn16','VVIn73','VVIp31','VVIp60','VVIq52','VVIv37','VVIv67']
MIG_EXTRA = ['VrZAG64','VrZAG83','VChr3a','VChr4a','VChr7b','VChr10b','VChr19a','VMC6E1','VMCNG4b9']
OLI_EXTRA = ['VVIn52','VVIn56','VVIn74','VVIp25b','VVIr09','VVIr21','VVIv36','VVip31','VVip77','VViq57']
RAN_EXTRA = ['EVA2','ISV2','ISV4','ISV3']

MIG_ALL = GENRES[:6] + ['VrZAG62'] + GENRES[6:] + MIG_EXTRA   # VrZAG64 not VVMD32 in Migliaro
MIG_ALL = ['VVS2','VVMD5','VVMD7','VVMD25','VVMD27','VVMD28','VrZAG62','VrZAG64','VrZAG79','VrZAG83',
           'VChr3a','VChr4a','VChr7b','VChr10b','VChr19a','VMC6E1','VMCNG4b9']
OLI_ALL = ['VVIn52','VVIn56','VVIn74','VVIp25b','VVIr09','VVIr21','VVIv36',
           'VVMD25','VVMD27','VVMD5','VVMD7','VVS2','VVip31','VVip77','VViq57','VrZAG62','VrZAG79']

# Master ordered locus list
_seen = set()
ALL_LOCI = []
for l in GENRES + LAC_EXTRA + MIG_EXTRA + OLI_EXTRA + RAN_EXTRA:
    if l not in _seen:
        ALL_LOCI.append(l)
        _seen.add(l)

RAN_ALL = GENRES + RAN_EXTRA
TEL_LOCI = ['VVS2','VVMD5','VVMD7','VVMD27','VVMD32','VrZAG62','VrZAG79']  # 7 of 9 GENRES081

# Which loci each source has
LOCUS_IN_SOURCE = {
    'VIVC (Ref)'   : set(GENRES),
    'Vassal/VIVC'  : set(GENRES),
    'Lacombe 2013' : set(GENRES) | set(LAC_EXTRA),
    'Migliaro 2019': set(MIG_ALL),
    'Oliveira 2020': set(OLI_ALL),
    'Rancho 2022'  : set(RAN_ALL),
    'Tello 2024'   : set(TEL_LOCI),
}

# ─── Helper: combined allele string A1/A2 sorted numerically ─────────────────
def make_allele(a1, a2):
    if pd.isna(a1) or pd.isna(a2): return None
    try:
        a1i, a2i = int(float(a1)), int(float(a2))
        return f"{min(a1i,a2i)}/{max(a1i,a2i)}"
    except:
        a1s, a2s = str(a1).strip(), str(a2).strip()
        return f"{a1s}/{a2s}" if a1s and a2s else None

# ─── Load VIVC reference (4 varieties, GENRES) ───────────────────────────────
vivc_raw = pd.read_csv(f'{DATA}/ssr_vivc.csv', low_memory=False)
vivc_raw = vivc_raw[vivc_raw['variety'] != 'Alleles:'].copy()
vivc_raw['variety'] = vivc_raw['variety'].str.strip().str.upper()
vivc_src = {}
for _, r in vivc_raw.iterrows():
    name = r['variety']
    for locus in GENRES:
        if locus in vivc_raw.columns:
            v = str(r[locus]).strip() if pd.notna(r[locus]) else None
            if v and v not in ('nan','None'):
                vivc_src.setdefault(name, {})[locus] = v

vassal_raw = pd.read_csv(f'{DATA}/ssr_vivc_vassal.csv', low_memory=False)
vassal_raw['variety'] = vassal_raw['variety'].str.strip().str.upper()
vassal_src = {}
for _, r in vassal_raw.iterrows():
    name = r['variety']
    for locus in GENRES:
        if locus in vassal_raw.columns:
            v = str(r[locus]).strip() if pd.notna(r[locus]) else None
            if v and v not in ('nan','None'):
                vassal_src.setdefault(name, {})[locus] = v

# ─── Load Lacombe 2013 (collision-resolved) ──────────────────────────────────
# Lacombe 2013 uses a different allele size standard vs VIVC.  Add offsets to
# get VIVC-equivalent bp sizes (confirmed on Cab Sauv, Pinot Noir, Chardonnay,
# Muscat a Petits Grains Blancs):
#   VVS2 +2, VVMD25 +1, VVMD27 +4, VVMD28 +1, VVMD32 +1; others: 0
# Note: VVMD5 offset is allele-size dependent (~+5 for alleles <232, ~+4 for ≥232).
# A +4 correction fixes the majority; ~25% of VVMD5 comparisons may still show 1bp discrepancy.
# Also: VVMD7 offset vs VIVC confirmed +1 on Rancho; Lacombe VVMD7 appears to match VIVC directly.
# VrZAG62/79 are NOT in Lacombe panel so no correction needed for those.
LAC_OFFSETS = {'VVS2': 2, 'VVMD5': 4, 'VVMD25': 1, 'VVMD27': 4, 'VVMD28': 1, 'VVMD32': 1}

def apply_lac_offset(allele_str, offset):
    """Apply +offset to each allele in 'a1/a2' string."""
    if not allele_str or offset == 0:
        return allele_str
    parts = allele_str.split('/')
    try:
        return '/'.join(str(int(p) + offset) for p in parts)
    except ValueError:
        return allele_str

lv_clean = pd.read_csv(f'{DATA}/ssr_lacombe_vivc.csv', low_memory=False)
lv_clean['vivc_prime_name'] = lv_clean['vivc_prime_name'].str.upper().str.strip()
LAC_ALL_LOCI = GENRES + LAC_EXTRA
lac_src = {}
for _, r in lv_clean.iterrows():
    name = str(r['vivc_prime_name']).strip()
    for locus in LAC_ALL_LOCI:
        col = f'ssr_{locus}'
        if col in lv_clean.columns:
            v = r.get(col)
            if pd.notna(v) and str(v).strip() not in ('nan', 'None', ''):
                raw = str(v).strip()
                offset = LAC_OFFSETS.get(locus, 0)
                lac_src.setdefault(name, {})[locus] = apply_lac_offset(raw, offset)

# ─── Load Migliaro 2019 ──────────────────────────────────────────────────────
mig_raw   = pd.read_csv(f'{DATA}/ssr_migliaro_2019.csv', low_memory=False)
mig_match = pd.read_csv(f'{DATA}/migliaro_vivc_match.csv', low_memory=False)
mig = mig_raw.merge(mig_match[['migliaro_id','vivc_prime_name']], on='migliaro_id', how='left')
mig['vivc_prime_name'] = mig['vivc_prime_name'].fillna(mig['variety_name']).str.upper()
mig_src = {}
for _, r in mig.iterrows():
    name = str(r['vivc_prime_name']).strip()
    for locus in MIG_ALL:
        a = make_allele(r.get(f'ssr_{locus}_a1'), r.get(f'ssr_{locus}_a2'))
        if a:
            mig_src.setdefault(name, {})[locus] = a

# ─── Load Oliveira 2020 ──────────────────────────────────────────────────────
oli_raw = pd.read_csv(f'{DATA}/ssr_oliveira_2020.csv', low_memory=False)
oli_raw['vivc_prime_name'] = oli_raw['vivc_match'].fillna(oli_raw['variety_name']).str.upper()
oli_src = {}
for _, r in oli_raw.iterrows():
    name = str(r['vivc_prime_name']).strip()
    for locus in OLI_ALL:
        a = make_allele(r.get(f'ssr_{locus}_a1'), r.get(f'ssr_{locus}_a2'))
        if a:
            oli_src.setdefault(name, {})[locus] = a

# ─── Load Rancho de la Merced 2022 ───────────────────────────────────────────
ran_raw = pd.read_csv(f'{DATA}/ssr_rancho_merced_vivc.csv', low_memory=False)
ran_raw['vivc_prime_name'] = ran_raw['vivc_prime_name'].fillna('').str.upper().str.strip()
ran_src = {}
for _, r in ran_raw.iterrows():
    name = str(r['vivc_prime_name']).strip()
    if not name:
        continue
    for locus in RAN_ALL:
        col = f'ssr_{locus}'
        if col in ran_raw.columns:
            v = r.get(col)
            if pd.notna(v) and str(v).strip() not in ('nan', 'None', ''):
                ran_src.setdefault(name, {})[locus] = str(v).strip()

# ─── Load Tello et al. 2024 (Serbian/Balkan varieties) ───────────────────────
# Alleles already VIVC-standardised in ssr_tello_2024.csv (offsets applied in build script)
tel_raw = pd.read_csv(f'{DATA}/ssr_tello_2024.csv', low_memory=False)
tel_raw['vivc_prime_name'] = tel_raw['vivc_prime_name'].fillna('').str.upper().str.strip()
tel_src = {}
for _, r in tel_raw.iterrows():
    name = str(r['vivc_prime_name']).strip()
    if not name:
        continue
    for locus in TEL_LOCI:
        col = f'ssr_{locus}'
        if col in tel_raw.columns:
            v = r.get(col)
            if pd.notna(v) and str(v).strip() not in ('nan', 'None', ''):
                tel_src.setdefault(name, {})[locus] = str(v).strip()

SOURCES_MAP = {
    'VIVC (Ref)'   : vivc_src,
    'Vassal/VIVC'  : vassal_src,
    'Lacombe 2013' : lac_src,
    'Migliaro 2019': mig_src,
    'Oliveira 2020': oli_src,
    'Rancho 2022'  : ran_src,
    'Tello 2024'   : tel_src,
}
SOURCE_LABELS = list(SOURCES_MAP.keys())
SOURCE_COLORS = [C['src_vivc'], C['src_vas'], C['src_lac'], C['src_mig'], C['src_oli'], C['src_ran'], C['src_tel']]

# All varieties across all sources
all_varieties = sorted(
    set(vivc_src) | set(vassal_src) | set(lac_src) | set(mig_src) | set(oli_src) | set(ran_src) | set(tel_src)
)
print(f'Total unique varieties: {len(all_varieties)}')

# ─── Build workbook ──────────────────────────────────────────────────────────
wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1 — Source Inventory & Tally
# ═══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Source Inventory"

# Title
ws1.merge_cells("A1:N1")
c = ws1['A1']
c.value = "SSR Molecular Marker Sources — Inventory & Tally"
c.font = Font(name="Arial", bold=True, size=14, color=C['hdr_bg'])
c.alignment = cal()
ws1.row_dimensions[1].height = 28

# Column headers
INV_HDRS = ["Source ID","Citation","Year","Marker Type","Platform / Array",
            "Loci (n)","Loci Names","N Varieties Total","N Varieties Obtained",
            "GENRES081 Overlap","Raw Data Available","Access Method","Notes","Status"]
ws1.row_dimensions[2].height = 36
for ci, h in enumerate(INV_HDRS, 1):
    cell = ws1.cell(row=2, column=ci, value=h)
    cell.font    = hfont(sz=9)
    cell.fill    = hfill(C['hdr_bg'])
    cell.alignment = cal(w=True)
    cell.border  = bdr

# Status colour map
STATUS_FILL = {
    "Full (locally available)": hfill("C6EFCE"),
    "Full (integrated)":        hfill("C6EFCE"),
    "Partial (reference varieties only)": hfill("FFEB9C"),
    "Not available (parentage results only)": hfill("FFC7CE"),
    "Not available (frequency data only)":    hfill("FFC7CE"),
    "Not available": hfill("FFC7CE"),
    "Not available (SSL/partner access barrier)": hfill("FFC7CE"),
}

# GENRES overlap per datasource
GENRES_SET = set(GENRES)
def genres_overlap(loci_str):
    if not isinstance(loci_str, str): return 0
    mentioned = set(re.findall(r'VV[A-Za-z0-9]+|VrZAG\d+', loci_str))
    return len(mentioned & GENRES_SET)

import re

for ri, row in datasrc.iterrows():
    er = ri + 3
    ws1.row_dimensions[er].height = 60
    bg = C['row_alt'] if ri % 2 else "FFFFFF"
    loci_n = len(str(row['loci']).split(',')) if isinstance(row['loci'], str) else '?'
    vals = [
        row['source_id'], row['citation'], str(row['year']),
        row['marker_type'], row['platform'],
        loci_n, row['loci'],
        row['n_varieties_total'] if pd.notna(row['n_varieties_total']) else '?',
        row['n_varieties_obtained'],
        genres_overlap(str(row['loci'])),
        "Yes" if str(row['raw_data_in_supp']).lower()=='true' else "No",
        row['access_method'], row['notes'], row['availability'],
    ]
    for ci, val in enumerate(vals, 1):
        cell = ws1.cell(row=er, column=ci, value=val)
        cell.font = bfont(sz=9)
        cell.alignment = cal(h="left", w=True)
        cell.fill = hfill(bg)
        cell.border = bdr
        if ci == 10:  # genres overlap
            cell.alignment = cal()
            if isinstance(val, int) and val >= 7:
                cell.fill = hfill(C['match'])
            elif isinstance(val, int) and val >= 4:
                cell.fill = hfill(C['single'])
        if ci == 11:  # raw data
            cell.alignment = cal()
            cell.fill = hfill(C['match']) if val == "Yes" else hfill(C['mismatch'])
        if ci == 14:  # status
            cell.fill = STATUS_FILL.get(str(val), hfill(bg))
            cell.font = bfont(sz=9, bold=True)

# Column widths
INV_WIDTHS = [18,42,6,18,30,8,50,12,12,12,10,32,55,28]
for i, w in enumerate(INV_WIDTHS, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.auto_filter.ref = f"A2:{get_column_letter(len(INV_HDRS))}2"
ws1.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2 — Locus × Source Coverage Matrix
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Locus Coverage Matrix")

ws2.merge_cells("A1:G1")
c = ws2['A1']
c.value = "SSR Locus Coverage — Which Markers Appear in Which Source"
c.font = Font(name="Arial", bold=True, size=13, color=C['hdr_bg'])
c.alignment = cal()
ws2.row_dimensions[1].height = 26

SRC_COLS = ["VIVC (Ref)", "Vassal/VIVC", "Lacombe 2013", "Migliaro 2019", "Oliveira 2020", "Rancho 2022", "Tello 2024"]
hdrs2 = ["Locus", "Panel Group"] + SRC_COLS + ["# Sources"]
for ci, h in enumerate(hdrs2, 1):
    cell = ws2.cell(row=2, column=ci, value=h)
    cell.font      = hfont(sz=10)
    cell.fill      = hfill(C['hdr_bg'])
    cell.alignment = cal()
    cell.border    = bdr

LOCUS_GROUP = {}
for l in GENRES:          LOCUS_GROUP[l] = "GENRES081 (9-locus)"
for l in LAC_EXTRA:       LOCUS_GROUP[l] = "Lacombe extra (20-panel)"
for l in MIG_EXTRA:       LOCUS_GROUP[l] = "Migliaro rootstock"
for l in OLI_EXTRA:       LOCUS_GROUP[l] = "Oliveira Iberian"
for l in RAN_EXTRA:       LOCUS_GROUP[l] = "Rancho extra (ISV/EVA)"

GROUP_COLORS = {
    "GENRES081 (9-locus)"       : "E3F2FD",
    "Lacombe extra (20-panel)"  : "F1F8E9",
    "Migliaro rootstock"        : "FFF8E1",
    "Oliveira Iberian"          : "FCE4EC",
    "Rancho extra (ISV/EVA)"    : "FFF3E0",
}

for ri, locus in enumerate(ALL_LOCI, 3):
    grp = LOCUS_GROUP.get(locus, "Other")
    bg  = GROUP_COLORS.get(grp, "FFFFFF")
    ws2.cell(row=ri, column=1, value=locus).font = bfont(bold=True, sz=10)
    ws2.cell(row=ri, column=1).fill = hfill(bg)
    ws2.cell(row=ri, column=1).alignment = cal(h="left")
    ws2.cell(row=ri, column=1).border = bdr
    ws2.cell(row=ri, column=2, value=grp).font = bfont(sz=9, color="444444")
    ws2.cell(row=ri, column=2).fill = hfill(bg)
    ws2.cell(row=ri, column=2).alignment = cal(h="left")
    ws2.cell(row=ri, column=2).border = bdr
    n_srcs = 0
    for si, src in enumerate(SRC_COLS, 3):
        has = locus in LOCUS_IN_SOURCE.get(src, set())
        cell = ws2.cell(row=ri, column=si)
        cell.value = "✓" if has else "—"
        cell.font  = Font(name="Arial", size=10, bold=has,
                          color=SOURCE_COLORS[si-3] if has else "BBBBBB")
        cell.fill  = hfill(C['match']) if has else hfill(C['absent'])
        cell.alignment = cal()
        cell.border = bdr
        if has: n_srcs += 1
    cnt_cell = ws2.cell(row=ri, column=len(hdrs2))
    cnt_cell.value = n_srcs
    cnt_cell.font  = bfont(bold=(n_srcs >= 3))
    cnt_cell.fill  = hfill(C['match']) if n_srcs >= 3 else (hfill(C['single']) if n_srcs >= 2 else hfill(bg))
    cnt_cell.alignment = cal()
    cnt_cell.border = bdr

ws2.column_dimensions['A'].width = 16
ws2.column_dimensions['B'].width = 26
for ci in range(3, len(hdrs2)+1):
    ws2.column_dimensions[get_column_letter(ci)].width = 16
ws2.freeze_panes = "A3"

# Add a variety count row at bottom
last_row = len(ALL_LOCI) + 3
ws2.cell(row=last_row, column=1, value="Varieties (n)").font = hfont(sz=9, color="000000")
ws2.cell(row=last_row, column=1).fill = hfill("E0E0E0")
ws2.cell(row=last_row, column=1).border = bdr
ws2.cell(row=last_row, column=2, value="Total accessions with ≥1 SSR locus").font = bfont(sz=9)
ws2.cell(row=last_row, column=2).border = bdr
var_counts = [len(vivc_src), len(vassal_src), len(lac_src), len(mig_src), len(oli_src), len(ran_src)]
for si, cnt in enumerate(var_counts, 3):
    c = ws2.cell(row=last_row, column=si, value=cnt)
    c.font = bfont(bold=True, sz=10)
    c.alignment = cal()
    c.border = bdr

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3 — SSR Profile Comparison (wide, variety × locus × source)
# ═══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("SSR Profile Comparison")

# For each locus, we show one column per source (only sources that have the locus)
# Column layout:
#   [VIVC Prime Name] [N Sources] [VVMD5: VIVC|Lac|Mig|Oli ...] [VVMD7: ...] ...

# Build column spec
LOCUS_SOURCE_COLS = []  # list of (locus, source_label)
for locus in ALL_LOCI:
    for src in SOURCE_LABELS:
        if locus in LOCUS_IN_SOURCE.get(src, set()):
            LOCUS_SOURCE_COLS.append((locus, src))

# Row 1: merged locus headers
# Row 2: source sub-headers
# Row 3+: data

ws3.merge_cells("A1:A2")
ws3['A1'].value = "VIVC Prime Name"
ws3['A1'].font  = hfont(sz=10)
ws3['A1'].fill  = hfill(C['hdr_bg'])
ws3['A1'].alignment = cal()
ws3['A1'].border = bdr

ws3.merge_cells("B1:B2")
ws3['B1'].value = "N Sources"
ws3['B1'].font  = hfont(sz=10)
ws3['B1'].fill  = hfill(C['hdr_bg'])
ws3['B1'].alignment = cal()
ws3['B1'].border = bdr

ws3.merge_cells("C1:C2")
ws3['C1'].value = "Mismatch Loci"
ws3['C1'].font  = hfont(sz=10)
ws3['C1'].fill  = hfill(C['hdr_bg'])
ws3['C1'].alignment = cal(w=True)
ws3['C1'].border = bdr

col_offset = 4  # D onwards

# Identify where each locus block starts
locus_blocks = {}
cur_col = col_offset
for locus in ALL_LOCI:
    srcs_for_locus = [s for s in SOURCE_LABELS if locus in LOCUS_IN_SOURCE.get(s, set())]
    if not srcs_for_locus:
        continue
    locus_blocks[locus] = (cur_col, srcs_for_locus)
    cur_col += len(srcs_for_locus)

GENRE_SET = set(GENRES)

# Row 1: locus merged headers
for locus, (start_col, srcs) in locus_blocks.items():
    end_col = start_col + len(srcs) - 1
    if end_col > start_col:
        ws3.merge_cells(start_row=1, start_column=start_col,
                        end_row=1, end_column=end_col)
    cell = ws3.cell(row=1, column=start_col, value=locus)
    cell.font      = hfont(sz=9, color="FFFFFF")
    cell.fill      = hfill(C['genre_hdr'] if locus in GENRE_SET else C['extra_hdr'])
    cell.alignment = cal()
    cell.border    = bdr

# Row 2: source sub-headers
for locus, (start_col, srcs) in locus_blocks.items():
    SRC_COL_COLORS = {
        'VIVC (Ref)'   : C['src_vivc'],
        'Vassal/VIVC'  : C['src_vas'],
        'Lacombe 2013' : C['src_lac'],
        'Migliaro 2019': C['src_mig'],
        'Oliveira 2020': C['src_oli'],
        'Rancho 2022'  : C['src_ran'],
        'Tello 2024'   : C['src_tel'],
    }
    for si, src in enumerate(srcs):
        col = start_col + si
        abbr = src.split()[0][:7]
        cell = ws3.cell(row=2, column=col, value=abbr)
        cell.font      = hfont(sz=8, color="FFFFFF")
        cell.fill      = hfill(SRC_COL_COLORS.get(src, "555555"))
        cell.alignment = cal()
        cell.border    = bdr

ws3.row_dimensions[1].height = 22
ws3.row_dimensions[2].height = 20

# Data rows
n_written = 0
for vi, variety in enumerate(all_varieties):
    row_idx = vi + 3
    bg = "FFFFFF" if vi % 2 == 0 else C['row_alt']

    # Count how many sources have data for this variety
    n_src = sum(1 for src in SOURCE_LABELS if variety in SOURCES_MAP[src])
    # Find mismatched loci
    mismatch_loci = []
    for locus in ALL_LOCI:
        vals = [SOURCES_MAP[src][variety][locus]
                for src in SOURCE_LABELS
                if variety in SOURCES_MAP[src] and locus in SOURCES_MAP[src][variety]]
        if len(vals) >= 2 and len(set(vals)) > 1:
            mismatch_loci.append(locus)

    # Name cell
    nc = ws3.cell(row=row_idx, column=1, value=variety)
    nc.font      = bfont(sz=9, bold=(n_src >= 2))
    nc.fill      = hfill(C['anchor'])
    nc.alignment = cal(h="left")
    nc.border    = bdr

    # N sources
    sc = ws3.cell(row=row_idx, column=2, value=n_src)
    sc.font      = bfont(sz=9, bold=True)
    sc.fill      = hfill(C['match']) if n_src >= 3 else (hfill(C['single']) if n_src >= 2 else hfill(bg))
    sc.alignment = cal()
    sc.border    = bdr

    # Mismatch loci list
    mc = ws3.cell(row=row_idx, column=3, value=", ".join(mismatch_loci) if mismatch_loci else "")
    mc.font      = bfont(sz=8, color="CC0000" if mismatch_loci else "888888")
    mc.fill      = hfill(C['mismatch']) if mismatch_loci else hfill(bg)
    mc.alignment = cal(h="left", w=True)
    mc.border    = bdr

    # Allele columns
    for locus, (start_col, srcs) in locus_blocks.items():
        locus_vals = []
        for si, src in enumerate(srcs):
            col = start_col + si
            allele = SOURCES_MAP[src].get(variety, {}).get(locus)
            cell   = ws3.cell(row=row_idx, column=col)
            if allele:
                cell.value = allele
                locus_vals.append(allele)
                cell.font  = bfont(sz=9)
                cell.alignment = cal()
                cell.fill  = hfill(bg)
            else:
                cell.value = ""
                cell.fill  = hfill(C['absent'])
                cell.font  = bfont(sz=8, color="CCCCCC")
                cell.alignment = cal()
            cell.border = bdr

        # Colour locus block: green=match, red=mismatch, amber=single
        unique_vals = set(locus_vals)
        if len(locus_vals) >= 2:
            block_fill = hfill(C['match']) if len(unique_vals) == 1 else hfill(C['mismatch'])
        elif len(locus_vals) == 1:
            block_fill = hfill(C['single'])
        else:
            block_fill = None

        if block_fill:
            for si in range(len(srcs)):
                col = start_col + si
                existing = ws3.cell(row=row_idx, column=col)
                if existing.value or not existing.value == "":
                    existing.fill = block_fill

    n_written += 1

# Freeze and column widths
ws3.freeze_panes = "D3"
ws3.column_dimensions['A'].width = 32
ws3.column_dimensions['B'].width = 9
ws3.column_dimensions['C'].width = 24
for col_letter_i in range(col_offset, cur_col):
    ws3.column_dimensions[get_column_letter(col_letter_i)].width = 10

ws3.row_dimensions[1].height = 22
ws3.row_dimensions[2].height = 18

print(f'Sheet 3: {n_written} varieties')

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4 — Mismatches Only
# ═══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("⚠ Cross-Source Mismatches")

ws4.merge_cells("A1:K1")
c = ws4['A1']
c.value = "Cross-Source Allele Mismatches — Varieties where sources disagree on allele calls"
c.font  = Font(name="Arial", bold=True, size=12, color="CC0000")
c.alignment = cal(h="left")
ws4.row_dimensions[1].height = 24

MM_HDRS = ["VIVC Prime Name","Locus","Panel Group",
           "VIVC (Ref)","Vassal/VIVC","Lacombe 2013","Migliaro 2019","Oliveira 2020","Rancho 2022","Tello 2024",
           "N Sources","Agree?","Notes"]
for ci, h in enumerate(MM_HDRS, 1):
    cell = ws4.cell(row=2, column=ci, value=h)
    cell.font      = hfont(sz=9)
    cell.fill      = hfill(C['hdr_bg'])
    cell.alignment = cal(w=True)
    cell.border    = bdr
ws4.row_dimensions[2].height = 30

mm_row = 3
for variety in all_varieties:
    for locus in ALL_LOCI:
        src_vals = {}
        for src in SOURCE_LABELS:
            if variety in SOURCES_MAP[src] and locus in SOURCES_MAP[src][variety]:
                src_vals[src] = SOURCES_MAP[src][variety][locus]
        if len(src_vals) < 2:
            continue  # need ≥2 sources to compare
        unique_vals = set(src_vals.values())
        agree = len(unique_vals) == 1

        # Only write mismatches
        if not agree:
            grp = LOCUS_GROUP.get(locus, "Other")
            note_parts = []
            for src, val in src_vals.items():
                note_parts.append(f"{src}: {val}")

            vals_row = [
                variety, locus, grp,
                src_vals.get('VIVC (Ref)', ""),
                src_vals.get('Vassal/VIVC', ""),
                src_vals.get('Lacombe 2013', ""),
                src_vals.get('Migliaro 2019', ""),
                src_vals.get('Oliveira 2020', ""),
                src_vals.get('Rancho 2022', ""),
                src_vals.get('Tello 2024', ""),
                len(src_vals),
                "✓" if agree else "⚠ MISMATCH",
                "; ".join(note_parts),
            ]
            ws4.row_dimensions[mm_row].height = 16
            for ci, val in enumerate(vals_row, 1):
                cell = ws4.cell(row=mm_row, column=ci, value=val)
                cell.font   = bfont(sz=9)
                cell.border = bdr
                cell.alignment = cal(h="left")
                if ci == 1:
                    cell.fill = hfill(C['anchor'])
                    cell.font = bfont(sz=9, bold=True)
                elif ci in [4, 5, 6, 7, 8, 9, 10]:
                    if val and val in src_vals.values():
                        all_srcs_vals = list(src_vals.values())
                        cell.fill = hfill(C['match']) if all_srcs_vals.count(val) == len(all_srcs_vals) else hfill(C['mismatch'])
                    else:
                        cell.fill = hfill(C['absent'])
                elif ci == 12:
                    cell.fill = hfill(C['match']) if agree else hfill(C['mismatch'])
                    cell.font = bfont(sz=9, bold=True, color="CC0000" if not agree else "1B5E20")
                else:
                    cell.fill = hfill("FFF8F8" if mm_row % 2 == 0 else "FFFFFF")
            mm_row += 1

# Also write the known VIVC reference discrepancies from crosssource_snippet
cs = pd.read_csv(f'{DATA}/ssr_crosssource_snippet.csv', low_memory=False)
cs_mismatches = cs[cs['VIVC_vs_Pub'] == '⚠'].copy()
if len(cs_mismatches):
    # Separator row
    ws4.cell(row=mm_row, column=1, value="── Known VIVC vs. Published Literature discrepancies (from crosssource_snippet.csv) ──")
    ws4.cell(row=mm_row, column=1).font = Font(name="Arial", bold=True, italic=True, size=10, color=C['hdr_bg'])
    ws4.merge_cells(start_row=mm_row, start_column=1, end_row=mm_row, end_column=13)
    ws4.row_dimensions[mm_row].height = 18
    mm_row += 1
    for _, r in cs_mismatches.iterrows():
        vals_row = [
            r['variety'], r['locus'], LOCUS_GROUP.get(r['locus'],'GENRES081 (9-locus)'),
            r['source_VIVC'], "", r['source_Published'], "", "", "", "",
            2, "⚠ MISMATCH",
            f"Source: {r['source_Pub_ref']}",
        ]
        for ci, val in enumerate(vals_row, 1):
            cell = ws4.cell(row=mm_row, column=ci, value=val)
            cell.font = bfont(sz=9)
            cell.border = bdr
            cell.alignment = cal(h="left")
            if ci == 1: cell.fill = hfill(C['anchor']); cell.font = bfont(sz=9, bold=True)
            elif ci == 4: cell.fill = hfill(C['mismatch'])  # VIVC val
            elif ci == 6: cell.fill = hfill(C['mismatch'])  # Published val
            elif ci == 12: cell.fill = hfill(C['mismatch']); cell.font = bfont(sz=9, bold=True, color="CC0000")
            else: cell.fill = hfill("FFFAF0")
        mm_row += 1

print(f'Sheet 4: {mm_row - 3} mismatch rows')

MM_WIDTHS = [32,12,24,13,13,13,13,13,13,13,9,14,50]
for i, w in enumerate(MM_WIDTHS, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.auto_filter.ref = f"A2:{get_column_letter(len(MM_HDRS))}2"
# Note: MM_HDRS now has 13 columns (added Rancho 2022, Tello 2024)
ws4.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 5 — Legend
# ═══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Legend")
ws5['A1'].value = "Legend & Colour Key"
ws5['A1'].font  = Font(name="Arial", bold=True, size=14, color=C['hdr_bg'])
ws5.row_dimensions[1].height = 28

legend_items = [
    (C['match'],    "FFFFFF", "✓  All sources agree (match)",
     "Allele calls are identical across all sources that have data for this locus"),
    (C['mismatch'], "CC0000", "⚠  Mismatch",
     "Two or more sources report different allele calls for the same variety × locus"),
    (C['single'],   "7B6200", "◑  Single-source only",
     "Only one source has data for this variety × locus — cannot cross-validate"),
    (C['absent'],   "999999", "—  Not in source",
     "This locus is not part of the marker panel used by this source"),
    (C['anchor'],   "4A235A", "Variety name column",
     "VIVC canonical prime name (anchor identity for all sources)"),
    ("1A237E",      "FFFFFF", "GENRES081 loci (dark blue header)",
     "9-locus OIV standard panel: VVS2, VVMD5, VVMD7, VVMD25, VVMD27, VVMD28, VVMD32, VrZAG62, VrZAG79"),
    ("4E342E",      "FFFFFF", "Extra loci (brown header)",
     "Additional loci used by Lacombe (20-panel), Migliaro (rootstock panel), Oliveira (Iberian panel), or Rancho 2022 (ISV/EVA)"),
]

for i, (bg, fg, label, desc) in enumerate(legend_items, 3):
    ws5.row_dimensions[i].height = 28
    c1 = ws5.cell(row=i, column=1, value=label)
    c1.font = Font(name="Arial", bold=True, size=11, color=fg)
    c1.fill = hfill(bg)
    c1.alignment = cal(h="left")
    c1.border = bdr
    c2 = ws5.cell(row=i, column=2, value=desc)
    c2.font = bfont(sz=11)
    c2.alignment = cal(h="left", w=True)
    c2.border = bdr
ws5.column_dimensions['A'].width = 36
ws5.column_dimensions['B'].width = 70

# Notes on VIVC being a compiled resource
ws5.row_dimensions[12].height = 16
ws5['A12'].value = "Note on VIVC SSR Data"
ws5['A12'].font  = Font(name="Arial", bold=True, size=12, color=C['hdr_bg'])
ws5.merge_cells("A12:B12")

note_text = (
    "The VIVC SSR profile database (~6,662 accessions) is a compiled resource: allele calls are collected by "
    "VIVC from multiple published studies using the GENRES081 9-locus standard panel. The 'VIVC (Ref)' column "
    "in this workbook reflects only the 4 reference varieties retrievable via the VIVC API endpoint "
    "(server returns fixed Cab Sauv, Chardonnay, Muscat, Pinot Noir regardless of query). The 'Vassal/VIVC' "
    "column shows the same 4 varieties as they appear in the Lacombe et al. 2013 supplement (MOESM7, the "
    "Vassal reference variety set), confirming that VIVC's reference profiles are drawn from Lacombe 2013.\n\n"
    "SIZE STANDARD CORRECTIONS (applied before comparison in this workbook):\n"
    "• Lacombe 2013 vs VIVC: VVS2 +2bp, VVMD5 +4bp*, VVMD25 +1bp, VVMD27 +4bp, VVMD28 +1bp, VVMD32 +1bp\n"
    "• Rancho 2022 vs VIVC (applied at source): VVS2 +1bp, VVMD5 +4bp, VVMD7 +1bp, VVMD27 +3bp, "
    "VVMD32 +1bp, VrZAG62 +2bp, VrZAG79 +3bp (confirmed on 4 ref. varieties)\n"
    "• Tello 2024 vs VIVC (applied at source): VVS2 −2bp, VVMD5 −4bp*, VVMD7 0bp, VVMD27 −4bp, "
    "VVMD32 −1bp, VrZAG62 0bp, VrZAG79 0bp (confirmed on 20+ overlap varieties; loci VVMD25/VVMD28 not genotyped)\n"
    "• Migliaro 2019 vs Oliveira 2020: 1bp offset at VVMD27, VrZAG62, VrZAG79 (Oliveira = Migliaro + 1bp)\n\n"
    "(*) VVMD5 in Lacombe has an allele-size-dependent offset: ~+5bp for alleles <232bp, ~+4bp for ≥232bp. "
    "Applying +4 correction leaves ~117 Lacombe×Rancho/Oliveira comparisons with 1bp residual "
    "discrepancy — these are NOT identity mismatches, just sizing artifacts. "
    "The genuinely unresolved mismatches involve: Migliaro vs Oliveira 1bp systematic differences, "
    "and a small number of variety-identity cases (different accessions with the same VIVC prime name)."
)
ws5['A13'].value = note_text
ws5['A13'].font  = bfont(sz=10)
ws5['A13'].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
ws5.merge_cells("A13:B13")
ws5.row_dimensions[13].height = 140

out = f'{BASE}/data/ssr_marker_profile_tally.xlsx'
wb.save(out)
print(f'\nSaved: {out}')
