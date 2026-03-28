"""
Extend constantini_2026_marker_support.csv using Sup. Table 8 pedigree
assignments and SNP dosage data (Sup. Table 6) from Constantini et al. 2026.

What this does
──────────────
Sup. Table 8 lists up to 84 parent-assignment rows for genotyped accessions.
After filtering to clean (non-ambiguous) assignments, this script:

  1. Identifies trios (child, P1, P2) not already in the existing file.
  2. For each new trio runs a Mendelian Exclusion Test on the SNP dosage calls
     (AA = hom-major, AB = het, BB = hom-minor) from Sup. Table 6.
  3. Computes a Dosage Consistency Score (DCS):

         DCS = 1 - (N_Mendelian_exclusions / N_informative_loci)

     At a homozygous locus (AA or BB) in a proposed parent, the child's dosage
     must include the parent's homozygous allele:
         parent=AA → child ∈ {AA, AB}      (child must have at least one A)
         parent=BB → child ∈ {AB, BB}      (child must have at least one B)

  4. Maps DCS → confidence_level and appends new rows to the existing CSV.

Run with:  python scripts/extend_constantini_support.py
"""

from __future__ import annotations

import re
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl

# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
BASE_DIR   = SCRIPT_DIR.parent
SUPP_DIR   = Path("/Users/brandonroy/Desktop/Informatique/Projects/VIVC App"
                  "/Supplemental Literature Data")

TABLE6_PATH = SUPP_DIR / "Data Sheet/Sup. Table 6.xlsx"
TABLE8_PATH = SUPP_DIR / "Data Sheet/Sup. Table 8.xlsx"
EXISTING_PATH = BASE_DIR / "data" / "constantini_2026_marker_support.csv"
ALIAS_PATH    = BASE_DIR / "data" / "marker_name_aliases.csv"
OUT_PATH      = BASE_DIR / "data" / "constantini_2026_marker_support.csv"   # overwrite in place

N_SNPS_TOTAL  = 10484
DOI           = "10.3389/fpls.2026.1771381"
STUDY_REF     = "Costantini et al. 2026"
CONFIRMED_YEAR = "2026"
MIN_INFORMATIVE = 100   # minimum homozygous-parent SNPs to score a pair

DCS_CONFIRMED = 0.97
DCS_PROBABLE  = 0.93
DCS_WEAK      = 0.85

# Patterns to reject from Parent 1 / Parent 2 cells in Table 8
_BAD_PARENT_RE = re.compile(
    r"unknown|different\s+proposition|mutation|false|"
    r"^\(|error|no\s+data|na\b",
    re.IGNORECASE,
)

# ── Helpers ───────────────────────────────────────────────────────────────────

def _clean_name(s) -> str:
    """Strip, uppercase; return '' for any NA-like or bracketed value."""
    s = str(s or "").strip()
    if not s or s.upper() in {"NAN", "NONE", "NA", "N/A", "ND", ""}:
        return ""
    if _BAD_PARENT_RE.search(s):
        return ""
    # Strip trailing parenthetical annotations like "NAME (comment)"
    s = re.sub(r"\s*\(.*\)\s*$", "", s).strip()
    return s.upper()


def load_alias_map() -> dict[str, str]:
    if not ALIAS_PATH.exists():
        return {}
    df = pd.read_csv(ALIAS_PATH, dtype=str).fillna("")
    if "alias" not in df.columns or "canonical_vivc_name" not in df.columns:
        return {}
    return {
        row["alias"].strip().upper(): row["canonical_vivc_name"].strip().upper()
        for _, row in df.iterrows()
        if row["alias"].strip()
    }


def load_table8() -> list[tuple[str, str, str]]:
    """
    Parse Sup. Table 8 → list of (child, parent1, parent2) triples.
    Skips rows with unknown / ambiguous / bracketed parent names.
    Returns names in UPPER case.
    """
    wb = openpyxl.load_workbook(TABLE8_PATH, read_only=True)
    ws = wb.active
    trios: list[tuple[str, str, str]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue                    # header rows
        child = _clean_name(row[1])    # col B = Genotype name
        p1    = _clean_name(row[2])    # col C = Parent 1
        p2    = _clean_name(row[3])    # col D = Parent 2
        if child and p1 and p2:
            trios.append((child, p1, p2))
    print(f"  Table 8 clean trios: {len(trios)}")
    return trios


def load_table6_genotypes(
    target_names: set[str],
) -> tuple[dict[str, int], dict[str, list[str]]]:
    """
    Parse Sup. Table 6 (SNP dosage matrix) for the target variety names.

    Returns
    -------
    snp_meta : {snp_id → row_index}  (informational)
    geno_matrix : {variety_name → [dosage_call, ...]}
        where dosage_call ∈ {'AA', 'AB', 'BB', 'NC'}
        length = number of SNP rows parsed
    """
    wb = openpyxl.load_workbook(TABLE6_PATH, read_only=True)
    ws = wb.active

    # Find column indices for target varieties
    target_cols: dict[int, str] = {}   # col_index → vivc_name
    geno_matrix: dict[str, list[str]] = {nm: [] for nm in target_names}
    snp_ids: list[str] = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 5:                      # "Sample_Name" row
            for j, val in enumerate(row):
                uval = str(val or "").strip().upper()
                if uval in target_names:
                    target_cols[j] = uval
        elif i > 6:
            # Data row: row[0]=None, row[1]=snp_name, row[2]=None, row[3]=metadata, ...
            snp_id = str(row[1] or row[0] or "").strip()
            if not snp_id:
                continue
            snp_ids.append(snp_id)
            for j, vname in target_cols.items():
                geno_matrix[vname].append(str(row[j] or "NC").strip())

    print(f"  Table 6 SNPs parsed: {len(snp_ids)}")
    print(f"  Varieties found: {list(target_cols.values())}")
    return {s: k for k, s in enumerate(snp_ids)}, geno_matrix


def compute_dcs(
    parent_genos: list[str],
    child_genos:  list[str],
) -> tuple[float | None, int, int]:
    """
    Dosage Consistency Score for a proposed parent–offspring pair.

    Dosage encoding: AA = hom-major, AB = het, BB = hom-minor, NC = missing.

    An informative locus is one where the parent is homozygous (AA or BB) and
    neither individual has a missing call.

    Returns (dcs, n_informative, n_excluded).
    Returns (None, ...) if n_informative < MIN_INFORMATIVE.
    """
    n_info = 0
    n_excl = 0
    for p, c in zip(parent_genos, child_genos):
        if p == "NC" or c == "NC":
            continue
        if p not in ("AA", "BB"):
            continue                    # het parent → not informative
        n_info += 1
        # parent=AA → child must have at least one A (AA or AB)
        # parent=BB → child must have at least one B (AB or BB)
        if p == "AA" and c == "BB":
            n_excl += 1
        elif p == "BB" and c == "AA":
            n_excl += 1
    if n_info < MIN_INFORMATIVE:
        return None, n_info, 0
    return 1.0 - n_excl / n_info, n_info, n_excl


def dcs_to_confidence(dcs: float, n_info: int) -> str:
    if dcs >= DCS_CONFIRMED and n_info >= 500:
        return "confirmed"
    if dcs >= DCS_PROBABLE:
        return "probable"
    if dcs >= DCS_WEAK:
        return "probable"
    return "disputed"


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    print("=== extend_constantini_support.py ===\n")

    alias_map = load_alias_map()

    # 1. Load existing rows to avoid duplicates
    print("[1/5] Loading existing Constantini marker support …")
    existing_df = pd.read_csv(EXISTING_PATH, dtype=str, keep_default_na=False)
    existing_pairs: set[tuple[str, str]] = set(
        zip(
            existing_df["child_variety"].str.strip().str.upper(),
            existing_df["parent_variety"].str.strip().str.upper(),
        )
    )
    print(f"  Existing rows: {len(existing_df)}; unique (child, parent) pairs: {len(existing_pairs)}")

    # 2. Load Table 8 trios
    print("\n[2/5] Parsing Sup. Table 8 …")
    raw_trios = load_table8()

    # Apply alias map
    def _resolve(name: str) -> str:
        return alias_map.get(name, name)

    trios = [
        (_resolve(c), _resolve(p1), _resolve(p2))
        for c, p1, p2 in raw_trios
    ]

    # Identify new trios
    new_trios = [
        (c, p1, p2) for c, p1, p2 in trios
        if (c, p1) not in existing_pairs and (c, p2) not in existing_pairs
    ]
    print(f"  New trios to score: {len(new_trios)}")
    if not new_trios:
        print("  Nothing new to add. Exiting.")
        return

    # 3. Load genotype data from Table 6 for all needed varieties
    print("\n[3/5] Loading genotype matrix from Sup. Table 6 …")
    all_needed: set[str] = set()
    for c, p1, p2 in new_trios:
        all_needed.update([c, p1, p2])

    _, geno_matrix = load_table6_genotypes(all_needed)
    available = {nm for nm, genos in geno_matrix.items() if genos}
    print(f"  Varieties with genotype data: {len(available)} / {len(all_needed)}")

    # 4. Score each new trio
    print("\n[4/5] Computing Dosage Consistency Scores …")
    new_rows: list[dict] = []
    by_conf: dict[str, int] = {}

    def _emit(child: str, parent: str, role: str) -> None:
        """Compute DCS for (child, parent) and append a row if scoreable."""
        pg = geno_matrix.get(parent, [])
        cg = geno_matrix.get(child,  [])
        if not pg or not cg:
            return
        dcs, n_info, n_excl = compute_dcs(pg, cg)
        if dcs is None:
            return
        conf = dcs_to_confidence(dcs, n_info)
        by_conf[conf] = by_conf.get(conf, 0) + 1

        # Chlorotype note from Table 8 if available (populated below)
        new_rows.append({
            "evidence_type":    "SNP array",
            "marker_type":      "Axiom Vitis22K",
            "n_markers":        N_SNPS_TOTAL,
            "lod_score":        np.nan,
            "study_reference":  STUDY_REF,
            "doi":              DOI,
            "confirmed_year":   CONFIRMED_YEAR,
            "confidence_level": conf,
            "notes": (
                f"DCS={dcs:.4f}; informative_loci={n_info}; exclusions={n_excl}"
            ),
            "child_variety":    child,
            "parent_variety":   parent,
            "parent_role":      role,
        })

    for child, p1, p2 in new_trios:
        if child not in available:
            continue
        if p1 in available:
            _emit(child, p1, "Parent 1")
        if p2 in available:
            _emit(child, p2, "Parent 2")

    print(f"\n  New rows added: {len(new_rows)}")
    for conf, n in sorted(by_conf.items()):
        print(f"    {conf:12s}: {n}")

    if not new_rows:
        print("  No new rows to write.")
        return

    # 5. Append and save
    print("\n[5/5] Writing extended CSV …")
    new_df  = pd.DataFrame(new_rows)
    combined = pd.concat([existing_df, new_df], ignore_index=True, sort=False)
    combined.to_csv(OUT_PATH, index=False)
    print(f"✓ Written {len(combined)} rows ({len(existing_df)} existing + {len(new_rows)} new) → {OUT_PATH}")


if __name__ == "__main__":
    main()
