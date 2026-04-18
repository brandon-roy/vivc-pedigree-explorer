"""
build_cunha_2020_pedigree.py
============================
Parse Supplementary Tables 3 (duos) and 5 (trios) from
Cunha et al. 2020 (Front. Plant Sci. 11:127) and write
data/cunha_2020_marker_support.csv in the standard
marker_support schema used across the project.

Marker support schema columns:
  child_variety, parent_1, parent_2, confidence_level,
  evidence_type, n_markers, lod_score, confirmed_year,
  study_reference, doi, notes

Novel relationships (References col empty) are tagged
"first_molecular_confirmation" in notes.
"""

import re
import unicodedata
import pathlib
import openpyxl
import pandas as pd

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR   = pathlib.Path(__file__).resolve().parents[1]
XL_PATH    = BASE_DIR / "data" / "Cunha et al 2020 Frontiers Plant Sci" / "Data Sheet 1.xlsx"
VIVC_PATH  = BASE_DIR / "data" / "master_accession_sheet.csv"
NMP_PATH   = BASE_DIR / "data" / "node_molecular_profile.csv"
OUT_PATH   = BASE_DIR / "data" / "cunha_2020_marker_support.csv"

DOI     = "10.3389/fpls.2020.00127"
YEAR    = 2020
STUDY   = "Cunha et al. 2020"
N_MARK  = 231   # SNP panel size

# ── Load VIVC prime names for name-based fallback matching ────────────────────
print("Loading VIVC name list…")
vivc_df  = pd.read_csv(VIVC_PATH, low_memory=False)
# Build normalised name → VIVC prime name lookup
def norm(s: str) -> str:
    """Upper-case, collapse whitespace, strip diacritics for fuzzy matching."""
    s = str(s).strip().upper()
    s = unicodedata.normalize("NFD", s)
    s = s.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", s)

vivc_names: dict[str, str] = {}
for _, row in vivc_df.iterrows():
    prime = str(row.get("vivc_prime_name") or "").strip()
    if not prime:
        continue
    vivc_names[norm(prime)] = prime

# Also build from node_molecular_profile for cross-check
nmp_df    = pd.read_csv(NMP_PATH)
nmp_names = set(nmp_df["vivc_prime_name"].dropna().str.strip().str.upper())

def resolve_name(cunha_name: str, icvv_map: dict) -> str:
    """Best-effort: ICVV lookup → VIVC prime match → fuzzy upper → raw."""
    s = str(cunha_name).strip()
    n = norm(s)
    # Exact VIVC name match
    if n in vivc_names:
        return vivc_names[n]
    # Return capitalised version that at least matches NMP
    if n in nmp_names:
        return s.upper()
    return s.title()


# ── Parse Excel ───────────────────────────────────────────────────────────────
print(f"Opening {XL_PATH.name}…")
wb = openpyxl.load_workbook(str(XL_PATH), read_only=True, data_only=True)

def safe_int(v):
    try: return int(v)
    except (TypeError, ValueError): return None

# ── SupplTable1: ICVV-SNP number → VIVC mapping ───────────────────────────────
ws1    = wb["SupplTable1"]
rows1  = list(ws1.iter_rows(values_only=True))
icvv_map: dict[int, dict] = {}
for row in rows1[4:]:
    if row[0] is None:
        break
    icvv_snp = safe_int(row[6])
    if icvv_snp is None:
        continue
    port_name  = str(row[1] or "").strip()
    vivc_no    = safe_int(row[3])
    vivc_prime = str(row[4] or "").strip()
    icvv_map[icvv_snp] = {
        "port_name":  port_name,
        "vivc_no":    vivc_no,
        "vivc_prime": vivc_prime if vivc_prime else port_name,
    }

# Explicit VIVC prime name overrides: norm(our_name) → correct VIVC prime name
# These fix cases where the Excel VIVC prime or accent-stripping diverges from VIVC usage.
VIVC_OVERRIDES: dict[str, str] = {
    "HEBEN":               "HEBIEN",              # VIVC prime is HEBIEN (41 uses in VIVC CSV)
    "ANADIA BRANCA":       "BRANCA DE ANADIA",    # VIVC prime has reversed word order
    "CASTALIA":            "CASTALIA",             # already correct
    "ALVARINHO LILAZ":     "ALVARINHO LILAS",      # Z vs S typo in Excel VIVC prime field
    "BRANCO GUIMARAES":    "BRANCO DE GUIMARAES", # missing 'de'
    "BRANCO DESCONHECIDO": "BRANCO DESCONHECIDO",  # keep as-is (no VIVC entry)
}

def icvv_name(icvv_no: int, cunha_name: str) -> str:
    """Return best VIVC prime name for an ICVV-SNP genotype number."""
    entry = icvv_map.get(icvv_no)
    if entry:
        prime = entry["vivc_prime"]
        resolved = resolve_name(prime, icvv_map)
    else:
        resolved = resolve_name(cunha_name, icvv_map)
    # Apply overrides keyed on the normalised resolved name
    return VIVC_OVERRIDES.get(norm(resolved), resolved)

# ── SupplTable5: Trios ────────────────────────────────────────────────────────
ws5    = wb["SupplTable5"]
rows5  = list(ws5.iter_rows(values_only=True))

out_rows = []
trios_count = 0
novel_count = 0

for row in rows5[4:]:
    if row[0] is None:
        break

    child_icvv = safe_int(row[0])
    child_name = str(row[2] or "").strip()
    p1_icvv    = safe_int(row[3])
    p1_name    = str(row[5] or "").strip()
    p2_icvv    = safe_int(row[6])
    p2_name    = str(row[8] or "").strip()
    lod        = row[10]
    ref_raw    = str(row[11] or "").strip()

    # Skip wild-vine accessions that have no VIVC cultivar name
    if "vsylvestris" in child_name.lower() or "#n/a" in child_name.lower():
        continue

    child_v = icvv_name(child_icvv, child_name)
    p1_v    = icvv_name(p1_icvv,    p1_name)
    p2_v    = icvv_name(p2_icvv,    p2_name)

    is_novel = (ref_raw == "")
    novel_flag = "first_molecular_confirmation; " if is_novel else ""
    if is_novel:
        novel_count += 1

    note = (f"{novel_flag}Cunha 2020 Front. Plant Sci. 11:127; "
            f"231 Vitis SNP markers; LOD={lod}")
    base = dict(
        evidence_type="SNP",
        marker_type="SNP",
        n_markers=N_MARK,
        lod_score=lod,
        study_reference=STUDY,
        doi=DOI,
        confirmed_year=YEAR,
        confidence_level="confirmed",
        notes=note,
    )
    out_rows.append({**base, "child_variety": child_v,
                     "parent_variety": p1_v, "parent_role": "Parent 1"})
    out_rows.append({**base, "child_variety": child_v,
                     "parent_variety": p2_v, "parent_role": "Parent 2"})
    trios_count += 1

print(f"Trios parsed: {trios_count}  (novel: {novel_count})")

# ── SupplTable3: Duos ─────────────────────────────────────────────────────────
ws3 = wb["SupplTable3"]
duos_count = 0

for i, row in enumerate(ws3.iter_rows(values_only=True)):
    if i < 4:
        continue
    if row[0] is None:
        break

    child_icvv = safe_int(row[0])
    child_name = str(row[2] or "").strip()
    p1_icvv    = safe_int(row[4])
    p1_name    = str(row[6] or "").strip()
    lod        = row[10]

    child_v = icvv_name(child_icvv, child_name)
    p1_v    = icvv_name(p1_icvv,    p1_name)

    out_rows.append({
        "child_variety":    child_v,
        "parent_variety":   p1_v,
        "parent_role":      "Parent 1",
        "confidence_level": "probable",
        "evidence_type":    "SNP",
        "marker_type":      "SNP",
        "n_markers":        N_MARK,
        "lod_score":        lod,
        "confirmed_year":   YEAR,
        "study_reference":  STUDY,
        "doi":              DOI,
        "notes":            f"Parent-offspring duo; Cunha 2020 Front. Plant Sci. 11:127; LOD={lod}",
    })
    duos_count += 1

print(f"Duos parsed:  {duos_count}")

# ── Write output ──────────────────────────────────────────────────────────────
df = pd.DataFrame(out_rows, columns=[
    "child_variety", "parent_variety", "parent_role",
    "evidence_type", "marker_type", "n_markers", "lod_score",
    "study_reference", "doi", "confirmed_year", "confidence_level", "notes",
])

df.to_csv(OUT_PATH, index=False)
print(f"\nWrote {len(df)} rows → {OUT_PATH.relative_to(BASE_DIR)}")

# ── Report: cross-check against VIVC Cunha rows ───────────────────────────────
vivc_ms = pd.read_csv(BASE_DIR / "data" / "vivc_confirmed_marker_support.csv", low_memory=False)

# VIVC is long-format: child_variety + parent_variety (one row per parent)
vivc_cunha = vivc_ms[vivc_ms["doi"].str.contains("00127", na=False)]
vivc_pairs = set()
for _, r in vivc_cunha.iterrows():
    vivc_pairs.add((norm(str(r["child_variety"])), norm(str(r["parent_variety"]))))

# Our CSV is long-format: child + parent_variety per row
cunha_pairs = set()
for _, r in df.iterrows():
    pv = str(r.get("parent_variety", "") or "").strip()
    if pv and pv.lower() not in ("nan", ""):
        cunha_pairs.add((norm(str(r["child_variety"])), norm(pv)))

overlap   = cunha_pairs & vivc_pairs
only_us   = cunha_pairs - vivc_pairs
only_vivc = vivc_pairs - cunha_pairs

print(f"\nCunha pairs total:    {len(cunha_pairs)}")
print(f"Also in VIVC:         {len(overlap)}")
print(f"Cunha-only (new):     {len(only_us)}")
print()
print("Novel trios (first molecular confirmation):")
seen_novel = set()
for r in out_rows:
    if "first_molecular_confirmation" in r.get("notes", ""):
        child = r["child_variety"]
        if child not in seen_novel:
            # Collect both parents for this child
            parents = [x["parent_variety"] for x in out_rows
                       if x["child_variety"] == child]
            print(f"  {child} × {' × '.join(parents[:2])}")
            seen_novel.add(child)
